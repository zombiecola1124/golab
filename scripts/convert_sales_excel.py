"""
GoLab v1.6.1 — 매출 Excel → sales_import.json 변환

입력: SALES/GLC 수익정리(2025)프로그램용.xlsx (월별 시트 1월~12월)
출력: web/data/sales_import.json

컬럼 매핑 (Row 5 헤더 기준):
  Col 0: Date       → saleDate
  Col 4: END USER   → vendor (고객)
  Col 5: 진행업체   → memo에 기록
  Col 6: Item       → itemName
  Col 7: Q'ty       → qty
  Col 8: 단가       → unitPrice (매출 단가)
  Col 9: 발주총액   → (검증용, unitPrice × qty)
  Col 10: 지출      → (참고: 원가)
  Col 11: 이익금    → (참고: 이익)
  Col 3: PO No.     → docNo
"""
import sys
import os
import io
import json
import hashlib
import re
from datetime import datetime, date

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

# ── 경로 계산 (크로스 플랫폼) ──
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
BASE_DIR = os.path.dirname(SCRIPT_DIR)
DEFAULT_INPUT = os.path.join(BASE_DIR, "SALES", "GLC 수익정리(2025)프로그램용.xlsx")
OUTPUT_DIR = os.path.join(BASE_DIR, "web", "data")
OUTPUT = os.path.join(OUTPUT_DIR, "sales_import.json")

# 월별 시트명 (데이터 시트만)
MONTH_SHEETS = ["1월", "2월", "3월", "4월", "5월", "6월",
                "7월", "8월", "9월", "10월 ", "11월", "12월"]
# 주의: "10월 " 뒤에 공백 포함 (원본 시트명 그대로)

HEADER_ROW = 5   # 1-indexed: Date, 발주처, 업종, PO No., END USER, ...
DATA_START = 6   # 데이터 시작 행


# ═══════════════════════════════════════════
# 정규화 함수 (convert_trade_excel.py 참조)
# ═══════════════════════════════════════════

def norm_str(v):
    """문자열 정규화: None → "", strip"""
    if v is None:
        return ""
    return str(v).strip()


def norm_upper(v):
    """대문자 정규화 (품번용)"""
    return norm_str(v).upper()


def norm_num(v):
    """숫자 정규화: 콤마/₩/공백 제거, 파싱 실패 → 0"""
    if v is None:
        return 0
    if isinstance(v, (int, float)):
        return v if v == v else 0  # NaN 체크
    s = str(v).replace(",", "").replace("₩", "").replace(" ", "").strip()
    try:
        return float(s) if "." in s else int(s)
    except (ValueError, TypeError):
        return 0


def norm_date(v):
    """날짜 정규화 → YYYY-MM-DD"""
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, date):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    # 엑셀 시리얼 넘버 (숫자)
    try:
        num = float(s)
        if 40000 < num < 50000:
            from datetime import timedelta
            base = datetime(1899, 12, 30)
            return (base + timedelta(days=int(num))).strftime("%Y-%m-%d")
    except (ValueError, TypeError):
        pass
    # YYYY-MM-DD, YYYY/MM/DD, YYYY.MM.DD
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(s[:10], fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    # datetime 문자열 "2025-01-02 00:00:00"
    try:
        return datetime.strptime(s[:19], "%Y-%m-%d %H:%M:%S").strftime("%Y-%m-%d")
    except ValueError:
        pass
    return s[:10] if len(s) >= 10 else ""


def make_idempotency_key(rec, sheet_name, row_num):
    """결정적 idempotency 키: sourceRowId 우선, hash fallback"""
    src = "|".join([
        norm_str(rec.get("saleDate", "")),
        norm_str(rec.get("vendor", "")),
        norm_str(rec.get("itemName", "")),
        str(norm_num(rec.get("qty", 0))),
        str(norm_num(rec.get("unitPrice", 0))),
        norm_str(rec.get("currency", "KRW")),
        norm_str(rec.get("docType", "SALE")),
        str(rec.get("sourceRowId", ""))
    ])
    h = hashlib.sha256(src.encode("utf-8")).hexdigest()[:16]
    return f"simp-{h}"


# ═══════════════════════════════════════════
# 메인 변환
# ═══════════════════════════════════════════

def process_sheet(ws, sheet_name):
    """월별 시트 1개 처리 → 레코드 리스트 반환"""
    records = []
    skipped = 0
    global_row = 0

    for row_idx, row in enumerate(ws.iter_rows(min_row=DATA_START, values_only=True), DATA_START):
        global_row = row_idx
        cells = list(row) if row else []

        # 최소 9개 컬럼 필요 (Date ~ 발주총액)
        if len(cells) < 10:
            cells.extend([None] * (10 - len(cells)))

        sale_date = norm_date(cells[0])
        vendor = norm_str(cells[4])        # END USER (고객)
        agent = norm_str(cells[5])         # 진행업체
        item_name = norm_str(cells[6])     # Item
        qty = norm_num(cells[7])
        unit_price = norm_num(cells[8])    # 매출 단가
        order_total = norm_num(cells[9])   # 발주총액 (검증용)
        cost = norm_num(cells[10]) if len(cells) > 10 else 0  # 지출
        profit = norm_num(cells[11]) if len(cells) > 11 else 0  # 이익금
        doc_no = norm_str(cells[3])        # PO No.

        # 빈 행 건너뛰기: 날짜 없거나 (vendor + itemName) 모두 비어있으면 스킵
        if not sale_date or (not vendor and not item_name):
            skipped += 1
            continue

        # sourceRowId: 시트명 + 행번호 (전역 고유)
        source_row_id = f"{sheet_name}!R{row_idx}"

        # 메모 조합: 진행업체 + PO No. (있을 경우)
        memo_parts = []
        if agent:
            memo_parts.append(f"진행:{agent}")
        if doc_no:
            memo_parts.append(f"PO:{doc_no}")
        memo = " / ".join(memo_parts)

        rec = {
            "saleDate": sale_date,
            "vendor": vendor,
            "itemCode": "",              # 이 엑셀에는 품번 컬럼 없음
            "itemName": item_name,
            "qty": qty,
            "unitPrice": int(unit_price) if unit_price == int(unit_price) else unit_price,
            "currency": "KRW",
            "vatIncluded": False,
            "priceBasis": "SUPPLY",
            "customerType": "B2B",
            "docType": "SALE",
            "memo": memo,
            "sourceRowId": source_row_id,
            "idempotencyKey": ""         # 아래에서 생성
        }

        rec["idempotencyKey"] = make_idempotency_key(rec, sheet_name, row_idx)
        records.append(rec)

    return records, skipped


def main():
    import openpyxl

    input_path = DEFAULT_INPUT

    # CLI 인자 처리
    if len(sys.argv) > 1:
        if sys.argv[1] == "--help":
            print("사용법: python convert_sales_excel.py [--file <경로>] [--test]")
            print(f"  기본 입력: {DEFAULT_INPUT}")
            print(f"  출력: {OUTPUT}")
            return
        if sys.argv[1] == "--file" and len(sys.argv) > 2:
            input_path = sys.argv[2]

    test_mode = "--test" in sys.argv

    # 입력 파일 확인
    if not os.path.isfile(input_path):
        print(f"[FATAL] 입력 파일 없음: {input_path}")
        sys.exit(1)

    print(f"=== GoLab 매출 Excel → JSON 변환 ===")
    print(f"입력: {input_path}")
    print(f"출력: {OUTPUT}\n")

    wb = openpyxl.load_workbook(input_path, read_only=True, data_only=True)

    all_records = []
    total_skipped = 0

    for sheet_name in MONTH_SHEETS:
        # 시트명 공백 처리 (10월 뒤 공백 등)
        actual_name = None
        for sn in wb.sheetnames:
            if sn.strip() == sheet_name.strip():
                actual_name = sn
                break

        if not actual_name:
            print(f"  [SKIP] 시트 '{sheet_name}' 없음")
            continue

        ws = wb[actual_name]
        records, skipped = process_sheet(ws, actual_name.strip())

        if records:
            print(f"  [{actual_name.strip():>3}] {len(records):>3}건 추출 (skip {skipped}건)")
            all_records.extend(records)
        else:
            print(f"  [{actual_name.strip():>3}]   0건 (skip {skipped}건)")

        total_skipped += skipped

    wb.close()

    # 테스트 모드: 최초 10건만
    if test_mode and len(all_records) > 10:
        all_records = all_records[:10]
        print(f"\n[TEST MODE] 10건만 출력")

    # 출력
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    with open(OUTPUT, "w", encoding="utf-8") as f:
        json.dump(all_records, f, ensure_ascii=False, indent=2)

    print(f"\n{'='*50}")
    print(f"총 {len(all_records)}건 변환 완료 → {OUTPUT}")
    print(f"스킵 {total_skipped}건 (빈 행/날짜 없음)")
    print(f"{'='*50}")

    # 요약 통계
    if all_records:
        vendors = set(r["vendor"] for r in all_records if r["vendor"])
        total_amount = sum(r["unitPrice"] * r["qty"] for r in all_records)
        dates = sorted(set(r["saleDate"] for r in all_records if r["saleDate"]))
        print(f"\n거래처: {len(vendors)}개")
        print(f"기간: {dates[0]} ~ {dates[-1]}")
        print(f"총 매출액: {total_amount:,.0f}원")

        # idempotencyKey 중복 체크
        keys = [r["idempotencyKey"] for r in all_records]
        dupes = len(keys) - len(set(keys))
        if dupes > 0:
            print(f"\n[WARN] idempotencyKey 중복: {dupes}건 — sourceRowId 확인 필요")
        else:
            print(f"[OK] idempotencyKey 중복 없음")


if __name__ == "__main__":
    main()
